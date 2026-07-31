"""
Module for generating textile waste analysis reports in multiple formats.
Supports PDF, Excel (XLSX), and CSV export.
"""

import io
import csv
from datetime import datetime
from typing import List, Optional, Dict, Any
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from database import WasteBatch, AnalysisResult, User


class ReportGenerator:
    """Generate reports for textile waste analysis data."""
    
    @staticmethod
    def generate_pdf_report(
        batches: List[WasteBatch],
        report_title: str = "Textile Waste Intelligence Report",
        user_name: str = None,
        date_range: str = None
    ) -> bytes:
        """
        Generate a PDF report of waste batch analysis.
        
        Args:
            batches: List of WasteBatch objects with analysis data
            report_title: Title for the report
            user_name: Name of user/organization
            date_range: Date range for the report
            
        Returns:
            PDF content as bytes
        """
        # Create PDF in memory
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1a2b4c'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2b5a8c'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        # Title and metadata
        elements.append(Paragraph(report_title, title_style))
        
        metadata_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        if user_name:
            metadata_text += f" | User: {user_name}"
        if date_range:
            metadata_text += f" | Period: {date_range}"
        
        elements.append(Paragraph(f"<i>{metadata_text}</i>", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary statistics
        elements.append(Paragraph("Executive Summary", heading_style))
        
        total_qty = sum(b.quantity for b in batches)
        analyzed_count = sum(1 for b in batches if b.analysis)
        avg_score = sum(b.analysis.overall_circularity_score for b in batches if b.analysis) / analyzed_count if analyzed_count > 0 else 0
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Batches', str(len(batches))],
            ['Total Weight (kg)', f'{total_qty:.2f}'],
            ['Analyzed Batches', str(analyzed_count)],
            ['Average Circularity Score', f'{avg_score:.2f}']
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b5a8c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Detailed batch information
        elements.append(Paragraph("Batch Details", heading_style))
        
        # Create table data
        batch_data = [['Batch ID', 'Fabric Type', 'Qty (kg)', 'Status', 'Circularity Score', 'Category']]
        
        for batch in batches:
            score = f"{batch.analysis.overall_circularity_score:.1f}" if batch.analysis else "N/A"
            category = batch.analysis.circularity_category if batch.analysis else "Not Analyzed"
            batch_data.append([
                str(batch.id),
                batch.fabric_type,
                f"{batch.quantity:.2f}",
                batch.status,
                score,
                category
            ])
        
        # Create table with proper column widths
        batch_table = Table(batch_data, colWidths=[0.8*inch, 1.5*inch, 1*inch, 1*inch, 1.5*inch, 1.7*inch])
        batch_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b5a8c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white])
        ]))
        
        elements.append(batch_table)
        
        # Build PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()
    
    @staticmethod
    def generate_excel_report(
        batches: List[WasteBatch],
        report_title: str = "Textile Waste Intelligence Report"
    ) -> bytes:
        """
        Generate an Excel (XLSX) report of waste batch analysis.
        
        Args:
            batches: List of WasteBatch objects with analysis data
            report_title: Title for the report
            
        Returns:
            Excel content as bytes
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Waste Analysis"
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 20
        worksheet.column_dimensions['I'].width = 12
        worksheet.column_dimensions['J'].width = 12
        worksheet.column_dimensions['K'].width = 12
        
        # Title
        worksheet['A1'] = report_title
        worksheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        worksheet['A1'].fill = PatternFill(start_color="1a2b4c", end_color="1a2b4c", fill_type="solid")
        worksheet.merge_cells('A1:K1')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Metadata
        worksheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet['A2'].font = Font(italic=True, size=10)
        
        # Headers
        headers = ['Batch ID', 'Fabric Type', 'Source', 'Qty (kg)', 'Color', 'Condition', 
                   'Status', 'Circularity Score', 'CO2 Saved (kg)', 'Water Saved (L)', 'Category']
        
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2b5a8c", end_color="2b5a8c", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        for row_idx, batch in enumerate(batches, start=5):
            worksheet.cell(row=row_idx, column=1).value = batch.id
            worksheet.cell(row=row_idx, column=2).value = batch.fabric_type
            worksheet.cell(row=row_idx, column=3).value = batch.source
            worksheet.cell(row=row_idx, column=4).value = batch.quantity
            worksheet.cell(row=row_idx, column=5).value = batch.color
            worksheet.cell(row=row_idx, column=6).value = batch.condition
            worksheet.cell(row=row_idx, column=7).value = batch.status
            
            if batch.analysis:
                worksheet.cell(row=row_idx, column=8).value = batch.analysis.overall_circularity_score
                worksheet.cell(row=row_idx, column=9).value = batch.analysis.co2_savings
                worksheet.cell(row=row_idx, column=10).value = batch.analysis.water_savings
                worksheet.cell(row=row_idx, column=11).value = batch.analysis.circularity_category
            else:
                worksheet.cell(row=row_idx, column=8).value = "N/A"
                worksheet.cell(row=row_idx, column=9).value = 0
                worksheet.cell(row=row_idx, column=10).value = 0
                worksheet.cell(row=row_idx, column=11).value = "Not Analyzed"
            
            # Format numbers
            for col in [4, 8, 9, 10]:
                cell = worksheet.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal='right')
                if col == 4:
                    cell.number_format = '0.00'
                elif col in [8, 9, 10]:
                    cell.number_format = '0.00'
            
            # Alternating row colors
            if row_idx % 2 == 0:
                for col in range(1, 12):
                    worksheet.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="f0f0f0", end_color="f0f0f0", fill_type="solid"
                    )
        
        # Summary section
        summary_row = len(batches) + 6
        worksheet.cell(row=summary_row, column=1).value = "SUMMARY"
        worksheet.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
        
        total_qty = sum(b.quantity for b in batches)
        total_co2 = sum(b.analysis.co2_savings for b in batches if b.analysis)
        total_water = sum(b.analysis.water_savings for b in batches if b.analysis)
        
        worksheet.cell(row=summary_row+1, column=1).value = "Total Weight (kg):"
        worksheet.cell(row=summary_row+1, column=2).value = total_qty
        
        worksheet.cell(row=summary_row+2, column=1).value = "Total CO2 Saved (kg):"
        worksheet.cell(row=summary_row+2, column=2).value = total_co2
        
        worksheet.cell(row=summary_row+3, column=1).value = "Total Water Saved (L):"
        worksheet.cell(row=summary_row+3, column=2).value = total_water
        
        # Convert to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    @staticmethod
    def generate_csv_report(batches: List[WasteBatch]) -> bytes:
        """
        Generate a CSV report of waste batch analysis.
        
        Args:
            batches: List of WasteBatch objects with analysis data
            
        Returns:
            CSV content as bytes
        """
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        
        # Write header
        headers = ['Batch ID', 'Fabric Type', 'Source', 'Quantity (kg)', 'Color', 'Condition',
                   'Status', 'Collection Date', 'Created At', 'Recyclability Score', 'Reuse Score',
                   'Sustainability Score', 'Material Recovery Score', 'Overall Circularity Score',
                   'Circularity Category', 'Recycling Strategy', 'CO2 Savings (kg)', 
                   'Water Savings (L)', 'Landfill Reduction (kg)']
        
        writer.writerow(headers)
        
        # Write data rows
        for batch in batches:
            row = [
                batch.id,
                batch.fabric_type,
                batch.source,
                f"{batch.quantity:.2f}",
                batch.color,
                batch.condition,
                batch.status,
                batch.collection_date,
                batch.created_at.isoformat() if batch.created_at else "",
            ]
            
            if batch.analysis:
                row.extend([
                    f"{batch.analysis.recyclability_score:.2f}",
                    f"{batch.analysis.reuse_score:.2f}",
                    f"{batch.analysis.sustainability_score:.2f}",
                    f"{batch.analysis.material_recovery_score:.2f}",
                    f"{batch.analysis.overall_circularity_score:.2f}",
                    batch.analysis.circularity_category,
                    batch.analysis.recycling_strategy,
                    f"{batch.analysis.co2_savings:.2f}",
                    f"{batch.analysis.water_savings:.2f}",
                    f"{batch.analysis.landfill_reduction:.2f}"
                ])
            else:
                row.extend(['N/A'] * 10)
            
            writer.writerow(row)
        
        # Add summary
        writer.writerow([])
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Batches', len(batches)])
        writer.writerow(['Total Weight (kg)', f"{sum(b.quantity for b in batches):.2f}"])
        writer.writerow(['Total CO2 Saved (kg)', f"{sum(b.analysis.co2_savings for b in batches if b.analysis):.2f}"])
        writer.writerow(['Total Water Saved (L)', f"{sum(b.analysis.water_savings for b in batches if b.analysis):.2f}"])
        
        # Convert to bytes
        csv_bytes = csv_buffer.getvalue().encode('utf-8')
        return csv_bytes
